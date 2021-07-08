import openpyxl
path="C:\code blocks\Documents\student_database_management.xlsx"
wb_obj=openpyxl.load_workbook(path)
sheet_obj = wb_obj.active
cell_obj=sheet_obj.cell(row=9,column=1)
print(cell_obj.value)
#printing one student complete record from excell sheet
for i in range(1,12):
    cell_obj = sheet_obj.cell(row=11,column=i)
    print(cell_obj.value)


import mysql.connector
mydb=mysql.connector.connect(host="localhost",user="root",password="Dharani@379",database="students")
mycursor=mydb.cursor()
print(mydb)
dbse=mydb.cursor()
dbse.execute("Show databases")
for i in dbse:
    print(i)

dbse=mydb.cursor()
dbse.execute("CREATE TABLE student12(roll_no int(10),regno int(10),Name Varchar(355),sem1 int(10),sem2 int(10),CGPA int(10),phn_num int,email_id varchar(355))")
dbse =mydb.cursor()
dbse.execute("Show tables")
for i in dbse:
    print(i)

cursor=mydb.cursor()
cursor.execute("Select * from student12")
for row in cursor:
    print(row)

import pandas as pd
df = pd.read_excel('student_database_managhement.xlsx')

import xlrd
import mysqldb
xl_sheet = xlrd.open_workbook("Student_database_management.xlsx")
xl_sheet
sheet_name=xl_sheet.sheet_names()
sheet_nsme

mydb=mysql.connector.connect(host="localhost",user="root",password="Dharani@379",database="students")
cursor=mydb.cursor()
for s in range(0,1):
    sheet=xl_shet.sheet_by_index(s)
    sql="INSERT INTO student1(roll_no,Reg_no,Name,sem1,sem2,CGPA,email_id)VALUES%s,%s,%s,%s,%s,%s,%s,%s)"
    for r in range(1,sheet.nrows):
        roll_no =sheet.cell(r,0).value
        reg_no = sheet.cell(r,1).value
        Name = sheet.cell(r,2).value
        sem1=sheet.cell(r,3).value
        sem2 = sheet.cell(r,4).value
        CGPA=sheet.cell(r,5).value
        email_id=sheet.cell(r,6).value
        values =(roll_no,reg_no,Name,sem1,sem2,CGPA,emai_id)
        cursor.execute(sql,values)
    mydb.commit()
    mycursor=mydb.cursor()
    mycursor.execute("Select * from student12")
    myresult=mycursor.fetchall()
    for x in myresult:
        print(x)

    mycursor=mydb.cursor()
    mycursor.execute("Select Name from students1 WHERE CGPA >6")
    myresult=mycursor.fetchall()
    for x in myrsult:
        print(x)

    mydb.commit()
    #closing th database connection
    mydb.close()
                             
